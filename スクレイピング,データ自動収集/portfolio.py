# -*- coding: utf-8 -*-
"""
矢島不動産（https://www.yajima-inc.net/）賃貸物件データ収集スクリプト
"""

from datetime import datetime
import re
import time
import urllib.parse

from bs4 import BeautifulSoup
import pandas as pd
import requests

"Exelの見た目の整理"
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://www.yajima-inc.net"
TARGET_URL = f"{BASE_URL}/result_building?type_id=1&mt=1&re_min=1&re_max=10000000000&pref[]=11&pa={{page}}"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
REQUEST_SLEEP_SEC = 1.5 


def clean_text(text: str) -> str:
    """余白や改行コードの整形"""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def parse_building_block(building_soup) -> list[dict]:
    """建物ブロックから部屋ごとのデータを抽出"""
    name_el = building_soup.find("div", class_="result_building_name_inner")
    building_name = name_el.contents[0].strip() if name_el and name_el.contents else ""

    # 所在地抽出
    address = ""
    span_tag = name_el.find("span") if name_el else None
    if span_tag:
        match = re.search(r"〇所在地[：:]\s*(.+?)\s*〇交通", span_tag.get_text(), re.S)
        if match:
            address = clean_text(match.group(1))

    table = building_soup.find_next_sibling("table")
    if not table:
        return []

    rooms = []
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue

        # 賃料
        rent_span = tds[3].find("span", class_="value")
        rent = f"{rent_span.get_text(strip=True)}万円" if rent_span else ""

        # 間取り
        layout_texts = list(tds[5].stripped_strings)
        layout = layout_texts[0] if layout_texts else ""

        # 詳細URL
        link = tds[6].find("a", class_="btn-push")
        detail_url = urllib.parse.urljoin(BASE_URL, link["href"]) if link and link.get("href") else ""

        rooms.append({
            "物件名称": building_name,
            "賃料": rent,
            "所在地": address,
            "間取り": layout,
            "詳細URL": detail_url,
        })

    return rooms

def extract_building_id(detail_url: str) -> str:
    match = re.search(r"s_bk_id=(\d+)", detail_url)
    return match.group(1) if match else detail_url

def get_full_address(session: requests.Session, detail_url: str) -> str | None:
    try:
        res = session.get(detail_url, timeout=10)
        res.raise_for_status()
        res.encoding = res.apparent_encoding
    except requests.RequestException as e:
        print(f"    Failed to fetch detail page: {e}")
        return None

    soup = BeautifulSoup(res.text, "html.parser")
    th_tag = soup.find("th", string=lambda t: t and "所在地" in t)
    if not th_tag:
        return None

    td_tag = th_tag.find_next("td")
    return clean_text(td_tag.get_text(strip=True)) if td_tag else None


def enrich_addresses(session: requests.Session, results: list[dict]) -> None:
    "正確な住所に置き換える。"
    cache: dict[str, str] = {}  

    print("Fetching accurate addresses from detail pages...")
    for i, room in enumerate(results, start=1):
        detail_url = room["詳細URL"]
        if not detail_url:
            continue
        
        building_id = extract_building_id(detail_url)
        
        if building_id not in cache:
            print(f"  [{i}/{len(results)}] {detail_url}")
            full_address = get_full_address(session, detail_url)
            cache[building_id] = full_address or room["所在地"]
            time.sleep(REQUEST_SLEEP_SEC)
        
        room["所在地"] = cache[building_id]

def fetch_all_properties(max_pages: int = 20) -> list[dict]:
    results = []
    
    with requests.Session() as session:
        session.headers.update(HEADERS)
        
        for page in range(1, max_pages + 1):
            url = TARGET_URL.format(page=page)
            print(f"Fetching page {page}...")
            
            try:
                res = session.get(url, timeout=10)
                res.raise_for_status()
                res.encoding = res.apparent_encoding
            except requests.RequestException as e:
                print(f"Request failed at page {page}: {e}")
                break

            soup = BeautifulSoup(res.text, "html.parser")
            buildings = soup.find_all("div", class_="result_building_wrap")
            
            if not buildings:
                print("No more items found. Finishing fetch.")
                break

            for b in buildings:
                results.extend(parse_building_block(b))
                
            time.sleep(1.5)
        if results:
            enrich_addresses(session, results)

    return results


def main():
    data = fetch_all_properties()
    if not data:
        print("No data collected.")
        return

    df = pd.DataFrame(data)
    df.drop_duplicates(subset=["詳細URL"], inplace=True)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"物件リスト_{today}.xlsx"
    
    # 1. Excelへ保存
    df.to_excel(filename, index=False, sheet_name="物件一覧")
    
    # 2. openpyxlでスタイル適用
    wb = openpyxl.load_workbook(filename)
    ws = wb["物件一覧"]

    # ヘッダーのデザイン設定
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # 落ち着いた紺色
    header_font = Font(color="FFFFFF", bold=True, name="メイリオ")                          # 白文字・太字
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # データ行の罫線,配置,フォントの設定
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    body_font = Font(name="メイリオ")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.font = body_font
            is_name_column = (cell.column == 1) # 1列目(A列)かどうか
            cell.alignment = Alignment(vertical="center", wrap_text=not is_name_column)

    # 列幅の自動調整
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # 全角は2文字分として長さを計算
            length = sum(2 if ord(c) > 256 else 1 for c in val_str)
            if length > max_len:
                max_len = length
        # 列ごとの幅調整
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = max_len + 4
        elif col_letter == 'C':
            ws.column_dimensions[col_letter].width = 35
        elif col_letter == 'E':
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 3. 上書き保存
    wb.save(filename)
    print(f"Saved {len(df)} rows to '{filename}'.")


if __name__ == "__main__":
    main()